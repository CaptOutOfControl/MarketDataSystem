"""
商超联盟门店数据上传系统
"""

import pandas as pd
import numpy as np
import socket
import threading
import time
import json
import os
import csv
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import traceback
import sys
import re

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# ====================== 编码修复函数 ======================
def fix_encoding(text):
    """尝试修复乱码文本"""
    if not text:
        return text
    
    if isinstance(text, str):
        # 检查是否包含乱码特征（中文字符被错误编码）
        # 判断标准：包含不可打印字符或编码异常
        try:
            # 如果文本可以正常编码解码，可能不需要修复
            text.encode('utf-8').decode('utf-8')
            return text
        except:
            try:
                # 先将文本用UTF-8编码（假设它当前是UTF-8字符串）
                utf8_bytes = text.encode('utf-8', errors='ignore')
                # 然后尝试用GBK解码
                fixed_text = utf8_bytes.decode('gbk', errors='ignore')
                return fixed_text
            except:
                # 如果失败，尝试另一种方式
                try:
                    # 有时候文本可能是GBK编码的字节被当作UTF-8字符串了
                    # 尝试先用latin1编码（不会丢失字节）
                    latin1_bytes = text.encode('latin1', errors='ignore')
                    # 然后用GBK解码
                    fixed_text = latin1_bytes.decode('gbk', errors='ignore')
                    return fixed_text
                except:
                    # 如果都失败，返回原文本
                    return text
    
    return text

# ====================== 辅助函数 ======================
def safe_int(value, default=0):
    """安全地将值转换为整数"""
    if pd.isna(value):
        return default
    
    try:
        # 如果是字符串，尝试提取数字
        if isinstance(value, str):
            # 移除空白字符
            value = value.strip()
            # 如果是空字符串，返回默认值
            if not value:
                return default
            # 尝试直接转换
            try:
                return int(value)
            except ValueError:
                # 尝试提取数字
                numbers = re.findall(r'\d+', value)
                if numbers:
                    return int(numbers[0])
                else:
                    return default
        # 如果是数字类型
        elif isinstance(value, (int, float, np.integer, np.floating)):
            return int(value)
        else:
            # 其他类型尝试转换
            return int(float(value))
    except Exception:
        return default

def safe_float(value, default=0.0):
    """安全地将值转换为浮点数"""
    if pd.isna(value):
        return default
    
    try:
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return default
            # 尝试提取数字（包括小数）
            numbers = re.findall(r'\d+\.?\d*', value)
            if numbers:
                return float(numbers[0])
            else:
                return default
        elif isinstance(value, (int, float, np.integer, np.floating)):
            return float(value)
        else:
            return float(value)
    except Exception:
        return default

# ====================== 数据解析器类 ======================
class DataParser:
    """数据解析器，用于解析不同格式的数据文件"""
    
    @staticmethod
    def parse_A(file_path):
        """解析A.xlsx文件"""
        try:
            df = pd.read_excel(file_path)
            retailer = "A"
            required_data = []
            
            for _, row in df.iterrows():
                prod_desc = row.get('prod_desc', '')
                if pd.isna(prod_desc):
                    continue
                    
                item = {
                    'retailer': retailer,
                    'prod_desc': str(prod_desc),
                    'price': safe_float(row.get('normal_price', 0)),
                    'unit': str(row.get('unit_dimemsion', '')) if not pd.isna(row.get('unit_dimemsion')) else '',
                    'quantity': safe_int(row.get('unit_number', 0)),
                    'supplier': str(row.get('vendor_name', '')) if not pd.isna(row.get('vendor_name')) else ''
                }
                required_data.append(item)
            
            print(f"从A.xlsx解析了 {len(required_data)} 条数据")
            return required_data
        except Exception as e:
            print(f"解析A.xlsx文件失败: {e}")
            traceback.print_exc()
            return []
    
    @staticmethod
    def parse_B(file_path):
        """解析B.csv文件"""
        try:
            df = pd.read_csv(file_path, encoding='utf-8-sig')
            retailer = "B"
            required_data = []
            
            for index, row in df.iterrows():
                try:
                    # 跳过空行
                    if pd.isna(row.iloc[0]) if len(row) > 0 else True:
                        continue
                    
                    prod_desc = row.get('prod_desc', '')
                    if pd.isna(prod_desc):
                        continue
                    
                    # 获取数量字段
                    quantity_val = row.get('unit_number', 0)
                    
                    item = {
                        'retailer': retailer,
                        'prod_desc': str(prod_desc),
                        'price': 0,  # B.csv中没有价格字段
                        'unit': str(row.get('unit_dimension', '')) if not pd.isna(row.get('unit_dimension')) else '',
                        'quantity': safe_int(quantity_val),
                        'supplier': str(row.get('vendor_name', '')) if not pd.isna(row.get('vendor_name')) else ''
                    }
                    required_data.append(item)
                    
                except Exception as e:
                    print(f"解析B.csv第{index}行时出错: {e}")
                    continue
            
            print(f"从B.csv解析了 {len(required_data)} 条数据")
            return required_data
        except Exception as e:
            print(f"解析B.csv文件失败: {e}")
            traceback.print_exc()
            return []
    
    @staticmethod
    def parse_C(file_path):
        """解析C.csv文件"""
        try:
            # 首先使用您的方法修复编码问题
            fixed_file = DataParser._fix_encoding_before_parse(file_path)
            
            # 然后尝试使用pandas读取修复后的文件
            try:
                # 尝试多种编码
                for encoding in ['utf-8-sig', 'gbk', 'utf-8', 'latin1']:
                    try:
                        df = pd.read_csv(fixed_file, encoding=encoding, low_memory=False, on_bad_lines='skip')
                        print(f"使用 {encoding} 编码成功读取修复后的C.csv，共 {len(df)} 行")
                        break
                    except:
                        continue
                else:
                    print("所有编码尝试都失败")
                    df = None
            except Exception as e:
                print(f"pandas读取失败: {e}")
                df = None
            
            if df is not None and len(df) > 0:
                retailer = "C"
                required_data = []
                
                desc_idx = 7 if len(df.columns) > 7 else -1
                price_idx = 11 if len(df.columns) > 11 else -1
                unit_idx = 18 if len(df.columns) > 18 else -1
                quantity_idx = 19 if len(df.columns) > 19 else -1
                supplier_idx = 23 if len(df.columns) > 23 else -1
                
                for index, row in df.iterrows():
                    try:
                        if len(row) == 0 or (isinstance(row.iloc[0], float) and np.isnan(row.iloc[0])):
                            continue
                        
                        prod_desc = str(row.iloc[desc_idx]) if desc_idx >= 0 and desc_idx < len(row) else ''
                        price_val = row.iloc[price_idx] if price_idx >= 0 and price_idx < len(row) else 0
                        unit = str(row.iloc[unit_idx]) if unit_idx >= 0 and unit_idx < len(row) else ''
                        quantity_val = row.iloc[quantity_idx] if quantity_idx >= 0 and quantity_idx < len(row) else 0
                        supplier = str(row.iloc[supplier_idx]) if supplier_idx >= 0 and supplier_idx < len(row) else ''
                        
                        if not prod_desc or prod_desc.strip() == '':
                            continue
                        
                        prod_desc = fix_encoding(prod_desc)
                        unit = fix_encoding(unit)
                        supplier = fix_encoding(supplier)
                        
                        try:
                            price = float(price_val) if not pd.isna(price_val) else 0
                        except:
                            price = 0
                        
                        try:
                            quantity = int(float(quantity_val)) if not pd.isna(quantity_val) else 0
                        except:
                            quantity = 0
                        
                        item = {
                            'retailer': retailer,
                            'prod_desc': prod_desc,
                            'price': safe_float(price_val),
                            'unit': unit,
                            'quantity': safe_int(quantity_val),
                            'supplier': supplier
                        }
                        required_data.append(item)
                        
                    except Exception as e:
                        continue
                
                print(f"从C.csv解析了 {len(required_data)} 条数据")
                return required_data
            else:
                return DataParser._parse_C_manually(file_path)
                
        except Exception as e:
            print(f"解析C.csv文件失败: {e}")
            traceback.print_exc()
            return DataParser._parse_C_manually(file_path)
    
    @staticmethod
    def _parse_C_manually(file_path):
        """手动解析C.csv文件"""
        retailer = "C"
        required_data = []
        
        try:
            # 使用您的方法修复编码
            fixed_file = DataParser._fix_encoding_before_parse(file_path)
            
            # 尝试不同编码读取修复后的文件
            encodings = ['utf-8-sig', 'gbk', 'latin1']
            content = None
            
            for encoding in encodings:
                try:
                    with open(fixed_file, 'r', encoding=encoding) as f:
                        content = f.read()
                    print(f"使用 {encoding} 编码读取修复后的文件成功")
                    break
                except UnicodeDecodeError:
                    continue
            
            if content is None:
                print("无法读取文件")
                return []
            
            lines = content.split('\n')
            
            for line in lines:
                if not line.strip():
                    continue
                
                fields = line.strip().split(',')
                
                if len(fields) < 24:
                    continue
                
                if fields[0] != 'letus':
                    continue
                
                prod_desc = fields[7] if len(fields) > 7 else ''
                price_str = fields[11] if len(fields) > 11 else '0'
                unit = fields[18] if len(fields) > 18 else ''
                quantity_str = fields[19] if len(fields) > 19 else '0'
                supplier = fields[23] if len(fields) > 23 else ''
                
                if not prod_desc or prod_desc.strip() == '':
                    continue
                
                prod_desc = fix_encoding(prod_desc)
                unit = fix_encoding(unit)
                supplier = fix_encoding(supplier)
                
                try:
                    price = float(price_str) if price_str.strip() != '' else 0
                except:
                    price = 0
                
                try:
                    quantity = int(float(quantity_str)) if quantity_str.strip() != '' else 0
                except:
                    quantity = 0
                
                item = {
                    'retailer': retailer,
                    'prod_desc': prod_desc,
                    'price': safe_float(price_str),
                    'unit': unit,
                    'quantity': safe_int(quantity_str),
                    'supplier': supplier
                }
                required_data.append(item)
            
            print(f"手动从C.csv解析了 {len(required_data)} 条数据")
            return required_data
            
        except Exception as e:
            print(f"手动解析C.csv失败: {e}")
            traceback.print_exc()
            return []
    
    @staticmethod
    def parse_D(file_path):
        """解析D.csv文件"""
        try:
            return DataParser.parse_C(file_path)
        except Exception as e:
            print(f"解析D.csv文件失败: {e}")
            traceback.print_exc()
            return []
    
    @staticmethod
    def parse_file(file_path):
        """根据文件类型调用相应的解析方法"""
        filename = os.path.basename(file_path)
        
        if filename.lower().endswith('.xlsx'):
            return DataParser.parse_A(file_path)
        elif filename.lower() == 'b.csv':
            return DataParser.parse_B(file_path)
        elif filename.lower() == 'c.csv':
            return DataParser.parse_C(file_path)
        elif filename.lower() == 'd.csv':
            return DataParser.parse_D(file_path)
        else:
            print(f"不支持的文件类型: {filename}")
            return []
        
    @staticmethod
    def _fix_encoding_before_parse(input_file):
        """
        在解析前修复编码问题 - 使用您的方法
        返回修复后的文件路径
        """
        import tempfile
        
        # 创建临时文件
        temp_dir = tempfile.gettempdir()
        base_name = os.path.basename(input_file)
        fixed_file = os.path.join(temp_dir, f"fixed_{base_name}")
        
        try:
            # 使用您的方法修复编码
            # 步骤1：以二进制方式读取文件，获取原始字节
            with open(input_file, 'rb') as f:
                raw_bytes = f.read()
            
            print(f"正在修复 {input_file} 的编码问题...")
            print(f"原始文件大小: {len(raw_bytes)} 字节")
            
            # 步骤2：尝试将字节用UTF-8解码（这会得到乱码字符串）
            try:
                mis_decoded_text = raw_bytes.decode('utf-8')
                print("检测到UTF-8错误编码")
            except UnicodeDecodeError:
                # 如果无法用UTF-8解码，可能是文件已经是GBK编码
                try:
                    # 尝试直接用GBK解码
                    fixed_text = raw_bytes.decode('gbk', errors='ignore')
                    print("文件已经是GBK编码，无需修复")
                    # 保存为GBK编码的文件
                    with open(fixed_file, 'w', encoding='gbk') as f:
                        f.write(fixed_text)
                    print(f"已保存修复后的文件: {fixed_file}")
                    return fixed_file
                except UnicodeDecodeError:
                    print("文件无法用GBK解码，尝试其他方法...")
                    # 如果无法用GBK解码，尝试用latin1（不会丢失字节）
                    fixed_text = raw_bytes.decode('latin1', errors='ignore')
                    print("使用latin1解码")
            
            # 步骤3：将乱码字符串用GBK编码（得到原始字节）
            # 注意：这里使用'ignore'参数跳过无法编码的字符
            original_gbk_bytes = mis_decoded_text.encode('gbk', errors='ignore')
            
            # 步骤4：将原始字节用GBK解码（得到正确文本）
            try:
                fixed_text = original_gbk_bytes.decode('gbk', errors='ignore')
                print("成功修复编码")
            except UnicodeDecodeError:
                print("无法用GBK解码原始字节，使用原始文本")
                fixed_text = mis_decoded_text
            
            # 步骤5：保存修复后的文件
            with open(fixed_file, 'w', encoding='gbk', newline='') as f:
                f.write(fixed_text)
            
            print(f"已成功修复编码并保存到: {fixed_file}")
            
            return fixed_file
            
        except Exception as e:
            print(f"修复编码时发生错误: {e}")
            import traceback
            traceback.print_exc()
            # 如果修复失败，返回原始文件
            return input_file

# ====================== 服务中心（服务端）类 ======================
class ServiceCenterServer:
    """服务中心（服务端）"""
    
    def __init__(self, host='127.0.0.1', control_port=5005, tcp_port=5006, udp_port=5007):
        self.host = host
        self.control_port = control_port
        self.tcp_port = tcp_port
        self.udp_port = udp_port
        
        self.received_data = {}
        self.receive_start_time = None
        self.receive_end_time = None
        
        self.output_file = 'service.xlsx'
        self.time_log_file = 'transmission_times.txt'
        
        self.lock = threading.Lock()
        
    def start(self):
        """启动服务中心"""
        print("启动服务中心...")
        
        control_thread = threading.Thread(target=self.start_udp_control_server)
        control_thread.daemon = True
        control_thread.start()
        
        print(f"UDP控制服务器监听在 {self.host}:{self.control_port}")
        print(f"TCP数据服务器监听在 {self.host}:{self.tcp_port}")
        print(f"UDP数据服务器监听在 {self.host}:{self.udp_port}")
        print("服务中心已启动，等待客户端连接...")
        print("按 Ctrl+C 停止服务中心")
        
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            print("\n服务中心关闭")
    
    def start_udp_control_server(self):
        """启动UDP控制服务器"""
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.bind((self.host, self.control_port))
        
        while True:
            try:
                data, addr = sock.recvfrom(1024)
                message = data.decode('utf-8')
                
                if message == "数据传输请求":
                    print(f"收到来自 {addr} 的数据传输请求")
                    
                    response = "请选择数据传输时的通信方式，基于TCP请选择1，基于UDP请选择0"
                    sock.sendto(response.encode('utf-8'), addr)
                    
                    data, addr = sock.recvfrom(1024)
                    choice = data.decode('utf-8')
                    
                    if choice in ['0', '1']:
                        if choice == '1':
                            print(f"客户端 {addr} 选择了TCP传输")
                            tcp_thread = threading.Thread(target=self.start_tcp_data_server, args=(addr,))
                            tcp_thread.daemon = True
                            tcp_thread.start()
                            
                            sock.sendto("我已准备完毕，请开始传输".encode('utf-8'), addr)
                        else:
                            print(f"客户端 {addr} 选择了UDP传输")
                            udp_thread = threading.Thread(target=self.start_udp_data_server, args=(addr,))
                            udp_thread.daemon = True
                            udp_thread.start()
                            
                            sock.sendto("我已准备完毕，请开始传输".encode('utf-8'), addr)
                    else:
                        print(f"无效的选择: {choice}")
                
            except Exception as e:
                print(f"UDP控制服务器错误: {e}")
                traceback.print_exc()
    
    def start_tcp_data_server(self, client_addr):
        """启动TCP数据服务器"""
        server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        server_socket.bind((self.host, self.tcp_port))
        server_socket.listen(1)
        
        server_socket.settimeout(30)
        
        try:
            conn, addr = server_socket.accept()
            
            if addr[0] == client_addr[0]:
                print(f"TCP数据连接已建立: {addr}")
                self.receive_data_tcp(conn, addr)
            else:
                print(f"拒绝来自 {addr} 的连接，期望 {client_addr}")
                conn.close()
        except socket.timeout:
            print(f"等待TCP连接超时")
        except Exception as e:
            print(f"TCP数据服务器错误: {e}")
            traceback.print_exc()
        finally:
            server_socket.close()
    
    def start_udp_data_server(self, client_addr):
        """启动UDP数据服务器"""
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.bind((self.host, self.udp_port))
        
        print(f"UDP数据服务器已启动，等待来自 {client_addr} 的数据")
        self.receive_data_udp(sock, client_addr)
    
    def receive_data_tcp(self, conn, addr):
        """通过TCP接收数据"""
        buffer = b""
        self.receive_start_time = time.time()
        data_list = []
        
        try:
            conn.settimeout(30)
            
            # 接收所有数据
            while True:
                chunk = conn.recv(4096)
                if not chunk:
                    break
                buffer += chunk
            
            print(f"从 {addr} 接收完成，共接收 {len(buffer)} 字节")
            
            if buffer:
                try:
                    # 解码为字符串
                    data_str = buffer.decode('utf-8')
                    
                    # 移除可能的结束消息
                    end_msg = "信息传输结束，我将断开连接"
                    if end_msg in data_str:
                        data_str = data_str.replace(end_msg, "")
                    
                    print(f"处理后的数据字符串长度: {len(data_str)}")
                    print(f"数据前500字符: {data_str[:500]}")
                    print(f"数据后500字符: {data_str[-500:]}")
                    
                    # 尝试解析为JSON
                    try:
                        # 尝试解析为单个JSON数组
                        items = json.loads(data_str)
                        if isinstance(items, list):
                            data_list.extend(items)
                            print(f"成功解析为JSON数组，包含 {len(items)} 条数据")
                        else:
                            data_list.append(items)
                            print(f"成功解析为JSON对象")
                    except json.JSONDecodeError as e:
                        print(f"JSON解析失败: {e}")
                        
                        # 尝试修复：查找所有JSON数组并合并
                        # 查找所有[...]模式
                        import re
                        json_pattern = r'\[.*?\]'
                        matches = re.findall(json_pattern, data_str, re.DOTALL)
                        
                        if matches:
                            print(f"找到 {len(matches)} 个可能的JSON数组")
                            for i, match in enumerate(matches):
                                try:
                                    items = json.loads(match)
                                    if isinstance(items, list):
                                        data_list.extend(items)
                                        print(f"第 {i+1} 个数组解析成功，包含 {len(items)} 条数据")
                                    else:
                                        data_list.append(items)
                                        print(f"第 {i+1} 个对象解析成功")
                                except:
                                    print(f"第 {i+1} 个匹配解析失败")
                        
                        # 如果正则匹配失败，尝试按批次分割
                        if not data_list:
                            # 查找明显的分隔符：][
                            if '][' in data_str:
                                parts = data_str.split('][')
                                print(f"按 '][' 分割为 {len(parts)} 部分")
                                
                                for i, part in enumerate(parts):
                                    # 修复每个部分，使其成为有效的JSON
                                    if i == 0:
                                        part = part + ']'
                                    elif i == len(parts) - 1:
                                        part = '[' + part
                                    else:
                                        part = '[' + part + ']'
                                    
                                    try:
                                        items = json.loads(part)
                                        if isinstance(items, list):
                                            data_list.extend(items)
                                            print(f"第 {i+1} 部分解析成功，包含 {len(items)} 条数据")
                                        else:
                                            data_list.append(items)
                                            print(f"第 {i+1} 部分解析成功")
                                    except:
                                        print(f"第 {i+1} 部分解析失败")
            
                except UnicodeDecodeError:
                    print("无法解码为UTF-8字符串")
            
        except socket.timeout:
            print(f"接收数据超时")
        except Exception as e:
            print(f"TCP数据接收错误: {e}")
            traceback.print_exc()
        finally:
            conn.close()
            
            self.receive_end_time = time.time()
            
            if data_list:
                print(f"准备保存 {len(data_list)} 条数据")
                self.save_received_data(data_list, addr)
            else:
                print("没有接收到有效数据")
    
    def receive_data_udp(self, sock, client_addr):
        """通过UDP接收数据"""
        self.receive_start_time = time.time()
        data_list = []
        packet_count = 0
        
        try:
            sock.settimeout(30)
            
            while True:
                try:
                    data, addr = sock.recvfrom(65535)
                    
                    if addr[0] != client_addr[0]:
                        continue
                    
                    message = data.decode('utf-8')
                    
                    if message == "信息传输结束，我将断开连接":
                        print(f"从 {addr} 接收到结束消息，共接收 {packet_count} 个数据包")
                        break
                    
                    try:
                        item = json.loads(message)
                        data_list.append(item)
                        packet_count += 1
                        
                        if packet_count % 100 == 0:
                            print(f"已接收 {packet_count} 个数据包")
                    except json.JSONDecodeError as e:
                        print(f"解析JSON数据失败: {e}")
                
                except socket.timeout:
                    print(f"接收数据超时")
                    break
            
        except Exception as e:
            print(f"UDP数据接收错误: {e}")
            traceback.print_exc()
        finally:
            sock.close()
            
            self.receive_end_time = time.time()
            
            if data_list:
                print(f"准备保存 {len(data_list)} 条数据")
                self.save_received_data(data_list, client_addr)
            else:
                print("没有接收到有效数据")
    
    def save_received_data(self, data_list, addr):
        """保存接收到的数据"""
        if not data_list:
            print("没有接收到数据")
            return
        
        receive_duration = self.receive_end_time - self.receive_start_time
        
        retailer_data = {}
        for item in data_list:
            retailer = item.get('retailer', 'Unknown')
            if retailer not in retailer_data:
                retailer_data[retailer] = []
            retailer_data[retailer].append(item)
        
        saved = self.save_to_excel(retailer_data)
        
        if saved:
            self.save_transmission_time(receive_duration, addr)
            
            print(f"数据已保存到 {self.output_file}")
            print(f"从 {addr} 接收数据耗时: {receive_duration:.4f} 秒")
            print(f"共保存 {len(data_list)} 条数据")
        else:
            print("保存数据失败")
    
    def save_to_excel(self, retailer_data):
        """将数据保存到Excel文件，每个零售商一个sheet"""
        try:
            with self.lock:
                if os.path.exists(self.output_file):
                    wb = openpyxl.load_workbook(self.output_file)
                else:
                    wb = Workbook()
                    if 'Sheet' in wb.sheetnames:
                        default_sheet = wb['Sheet']
                        wb.remove(default_sheet)
                
                for retailer, data in retailer_data.items():
                    sheet_name = str(retailer)[:31]
                    
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        if ws.max_row > 1:
                            ws.delete_rows(2, ws.max_row - 1)
                    else:
                        ws = wb.create_sheet(title=sheet_name)
                        headers = ['经销商', '产品描述', '价格', '商品单位', '商品数量', '供货商']
                        ws.append(headers)
                    
                    for item in data:
                        row = [
                            item.get('retailer', ''),
                            item.get('prod_desc', ''),
                            item.get('price', 0),
                            item.get('unit', ''),
                            item.get('quantity', 0),
                            item.get('supplier', '')
                        ]
                        ws.append(row)
                    
                    print(f"为零售商 {retailer} 添加了 {len(data)} 条数据到sheet '{sheet_name}'")
                
                wb.save(self.output_file)
                return True
            
        except Exception as e:
            print(f"保存Excel文件失败: {e}")
            traceback.print_exc()
            return self.save_to_csv(retailer_data)
    
    def save_to_csv(self, retailer_data):
        """将数据保存为CSV文件（备用方案）"""
        try:
            all_data = []
            for retailer, data in retailer_data.items():
                for item in data:
                    item['retailer'] = retailer
                    all_data.append(item)
            
            df = pd.DataFrame(all_data)
            df.to_csv('service.csv', index=False, encoding='utf-8')
            print(f"数据已保存到 service.csv")
            return True
        except Exception as e:
            print(f"保存CSV文件失败: {e}")
            traceback.print_exc()
            return False
    
    def save_transmission_time(self, duration, addr):
        """保存传输时间到文件"""
        try:
            with open(self.time_log_file, 'a', encoding='utf-8') as f:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                f.write(f"{timestamp} - 从 {addr} 接收耗时: {duration:.4f}秒\n")
            print(f"传输时间已记录到 {self.time_log_file}")
        except Exception as e:
            print(f"保存传输时间失败: {e}")

# ====================== 门店（客户端）类 ======================
class StoreClient:
    """门店（客户端）"""
    
    def __init__(self, host='127.0.0.1', control_port=5005, tcp_port=5006, udp_port=5007):
        self.host = host
        self.control_port = control_port
        self.tcp_port = tcp_port
        self.udp_port = udp_port
        
        self.transmit_start_time = None
        self.transmit_end_time = None
        
    def connect_to_service_center(self, file_path, protocol_choice='1'):
        """连接到服务中心并传输数据"""
        print(f"连接到服务中心，传输文件: {file_path}")
        
        if not os.path.exists(file_path):
            print(f"文件 {file_path} 不存在")
            return
        
        data = DataParser.parse_file(file_path)
        
        if not data:
            print("没有解析到数据，无法传输")
            return
        
        print(f"成功解析 {len(data)} 条数据")
        
        print("发送数据传输请求...")
        control_sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        
        try:
            control_sock.sendto("数据传输请求".encode('utf-8'), (self.host, self.control_port))
            
            control_sock.settimeout(10)
            response, _ = control_sock.recvfrom(1024)
            response_text = response.decode('utf-8')
            print(f"收到回复: {response_text}")
            
            if protocol_choice in ['0', '1']:
                control_sock.sendto(protocol_choice.encode('utf-8'), (self.host, self.control_port))
                
                ready_msg, _ = control_sock.recvfrom(1024)
                ready_text = ready_msg.decode('utf-8')
                print(f"收到回复: {ready_text}")
                
                if protocol_choice == '1':
                    self.transmit_data_tcp(data)
                else:
                    self.transmit_data_udp(data)
            else:
                print("无效的协议选择")
                
        except socket.timeout:
            print("连接超时")
        except Exception as e:
            print(f"连接错误: {e}")
            traceback.print_exc()
        finally:
            control_sock.close()
    
    def transmit_data_tcp(self, data):
        """通过TCP传输数据"""
        print("开始通过TCP传输数据...")
        
        self.transmit_start_time = time.time()
        
        sock = None
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(30)
            sock.connect((self.host, self.tcp_port))
            
            # 关键修改：将所有数据作为一个JSON数组发送
            json_data = json.dumps(data, ensure_ascii=False).encode('utf-8')
            
            # 先发送数据长度（可选，帮助服务端知道数据大小）
            data_len = len(json_data)
            print(f"发送数据长度: {data_len} 字节")
            
            # 发送数据
            sock.sendall(json_data)
            print(f"已发送 {len(data)} 条数据")
            
            # 发送结束消息
            end_msg = "信息传输结束，我将断开连接"
            sock.sendall(end_msg.encode('utf-8'))
            
            self.transmit_end_time = time.time()
            transmit_duration = self.transmit_end_time - self.transmit_start_time
            
            print(f"数据传输完成，耗时: {transmit_duration:.4f} 秒")
            
            self.save_transmission_time(transmit_duration)
            
        except socket.timeout:
            print("传输数据超时")
        except Exception as e:
            print(f"TCP传输错误: {e}")
            traceback.print_exc()
        finally:
            if sock:
                try:
                    sock.close()
                except:
                    pass
    
    def transmit_data_udp(self, data):
        """通过UDP传输数据"""
        print("开始通过UDP传输数据...")
        
        self.transmit_start_time = time.time()
        
        sock = None
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            sock.settimeout(10)
            
            for i, item in enumerate(data):
                json_str = json.dumps(item, ensure_ascii=False)
                sock.sendto(json_str.encode('utf-8'), (self.host, self.udp_port))
                
                if (i + 1) % 100 == 0:
                    print(f"已发送 {i + 1}/{len(data)} 条数据")
                
                time.sleep(0.001)
            
            end_msg = "信息传输结束，我将断开连接"
            sock.sendto(end_msg.encode('utf-8'), (self.host, self.udp_port))
            
            self.transmit_end_time = time.time()
            transmit_duration = self.transmit_end_time - self.transmit_start_time
            
            print(f"数据传输完成，耗时: {transmit_duration:.4f} 秒")
            
            self.save_transmission_time(transmit_duration)
            
        except socket.timeout:
            print("传输数据超时")
        except Exception as e:
            print(f"UDP传输错误: {e}")
            traceback.print_exc()
        finally:
            if sock:
                sock.close()
    
    def save_transmission_time(self, duration):
        """保存传输时间到文件"""
        try:
            with open('client_transmission_time.txt', 'a', encoding='utf-8') as f:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                f.write(f"{timestamp} - 传输耗时: {duration:.4f}秒\n")
            print(f"传输时间已记录到 client_transmission_time.txt")
        except Exception as e:
            print(f"保存传输时间失败: {e}")

# ====================== 主函数 ======================
def main():
    """主函数，演示系统使用"""
    
    print("=" * 60)
    print("门店数据上传系统")
    print("=" * 60)
    
    mode = input("请选择模式 (1-服务中心, 2-门店): ")
    
    if mode == '1':
        server = ServiceCenterServer()
        server.start()
        
    elif mode == '2':
        client = StoreClient()
        
        print("\n请选择要传输的文件:")
        print("1. A.xlsx")
        print("2. B.csv")
        print("3. C.csv")
        print("4. D.csv")
        
        file_choice = input("请输入文件编号 (1-4): ")
        protocol_choice = input("请选择传输协议 (1-TCP, 0-UDP): ")
        
        files = {
            '1': 'A.xlsx',
            '2': 'B.csv',
            '3': 'C.csv',
            '4': 'D.csv'
        }
        
        if file_choice in files:
            file_path = files[file_choice]
            client.connect_to_service_center(file_path, protocol_choice)
        else:
            print("无效的文件选择")
    else:
        print("无效的模式选择")

if __name__ == "__main__":
    main()